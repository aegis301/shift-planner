import csv
from dataclasses import dataclass
from datetime import date
from io import BytesIO, StringIO
from typing import Any

from sqlalchemy.orm import Session

from app.services.export_colors import MemberPastelPalette, member_pastel_palette
from app.services.matrix import get_planning_matrix
from app.services.roster_matrix import get_roster_matrix
from app.services.team_members import planning_display_name


@dataclass(frozen=True)
class RosterExportColumn:
    key: str
    title: str


@dataclass(frozen=True)
class RosterExportCell:
    member_name: str
    palette: MemberPastelPalette | None = None


@dataclass(frozen=True)
class RosterExportRow:
    day: date
    weekday: str
    cells: list[RosterExportCell | None]


@dataclass(frozen=True)
class RosterExportTable:
    period_label: str
    columns: list[RosterExportColumn]
    rows: list[RosterExportRow]


def _team_member_label(member: Any) -> str:
    return f"{member.first_name} {member.last_name}".strip()


def _planning_member_label(member: Any) -> str:
    return planning_display_name(
        nickname=getattr(member, "nickname", None),
        last_name=getattr(member, "last_name", ""),
    )


def _period_label(year: int, month: int) -> str:
    return f"{year}-{month:02d}"


def _slot_time_label(slot: Any) -> str:
    if slot.starts_at is None or slot.ends_at is None:
        return ""
    return f"{slot.starts_at.strftime('%H:%M')}-{slot.ends_at.strftime('%H:%M')}"


def _slot_column_title(slot: Any) -> str:
    base = slot.template_code or slot.label or "Slot"
    if slot.variant_label:
        base = f"{base} · {slot.variant_label}"
    time_label = _slot_time_label(slot)
    if time_label:
        base = f"{base} ({time_label})"
    return f"{base} #{slot.position}"


def _template_column_title(template: Any, fallback_slot: Any | None = None) -> str:
    if template is not None:
        code = getattr(template, "code", None)
        name = getattr(template, "name", None)
        if code and name:
            return f"{code} - {name}"
        if code:
            return code
        if name:
            return name
    if fallback_slot is None:
        return "Shift type"
    return fallback_slot.template_code or fallback_slot.label or "Shift type"


def _weekday_short(weekday: str) -> str:
    mapping = {
        "Monday": "Mo",
        "Tuesday": "Di",
        "Wednesday": "Mi",
        "Thursday": "Do",
        "Friday": "Fr",
        "Saturday": "Sa",
        "Sunday": "So",
    }
    if weekday in mapping:
        return mapping[weekday]
    if len(weekday) >= 2:
        return weekday[:2]
    return weekday


def build_roster_export_table(
    db: Session, planning_period_id: int, *, organization_id: int, shift_group_id: int | None = None
) -> RosterExportTable:
    matrix = get_roster_matrix(db, planning_period_id, organization_id=organization_id, shift_group_id=shift_group_id)
    slots_sorted = sorted(
        matrix.slots,
        key=lambda slot: (
            slot.template_code or "",
            slot.shift_template_id or 0,
            slot.shift_variant_id or 0,
            slot.position,
            slot.label or "",
            slot.id,
        ),
    )
    slot_to_column: dict[int, int] = {}
    column_index: dict[tuple[int | None, int | None, int, str | None], int] = {}
    columns: list[RosterExportColumn] = []
    for slot in slots_sorted:
        signature = (
            slot.shift_template_id,
            slot.shift_variant_id,
            slot.position,
            slot.label,
        )
        idx = column_index.get(signature)
        if idx is None:
            idx = len(columns)
            column_index[signature] = idx
            columns.append(RosterExportColumn(key=str(idx), title=_slot_column_title(slot)))
        slot_to_column[slot.id] = idx

    assignments = {assignment.roster_slot_id: assignment for assignment in matrix.assignments}
    members = {member.id: member for member in matrix.team_members}
    slot_for_day_column = {(slot.slot_date, slot_to_column[slot.id]): slot for slot in matrix.slots}

    rows: list[RosterExportRow] = []
    for day in matrix.days:
        row_cells: list[RosterExportCell | None] = []
        for col_idx in range(len(columns)):
            slot = slot_for_day_column.get((day.date, col_idx))
            if slot is None:
                row_cells.append(None)
                continue
            assignment = assignments.get(slot.id)
            if assignment is None:
                row_cells.append(None)
                continue
            member = members.get(assignment.team_member_id)
            if member is None:
                row_cells.append(None)
                continue
            row_cells.append(
                RosterExportCell(
                    member_name=_planning_member_label(member),
                    palette=member_pastel_palette(member.id),
                )
            )
        rows.append(RosterExportRow(day=day.date, weekday=_weekday_short(day.weekday), cells=row_cells))

    return RosterExportTable(
        period_label=_period_label(matrix.planning_period.year, matrix.planning_period.month),
        columns=columns,
        rows=rows,
    )


def build_roster_export_table_by_template(
    db: Session, planning_period_id: int, *, organization_id: int, shift_group_id: int | None = None
) -> RosterExportTable:
    matrix = get_roster_matrix(db, planning_period_id, organization_id=organization_id, shift_group_id=shift_group_id)
    template_by_id = {template.id: template for template in matrix.shift_templates}
    columns: list[RosterExportColumn] = []
    column_index: dict[tuple[int | None, str | None, str | None], int] = {}
    for template in matrix.shift_templates:
        signature = (template.id, template.code, None)
        column_index[signature] = len(columns)
        columns.append(
            RosterExportColumn(
                key=f"template-{template.id}",
                title=_template_column_title(template),
            )
        )
    slots_sorted = sorted(
        matrix.slots,
        key=lambda slot: (
            slot.template_code or "",
            slot.shift_template_id or 0,
            slot.position,
            slot.shift_variant_id or 0,
            slot.id,
        ),
    )
    slot_to_column: dict[int, int] = {}
    for slot in slots_sorted:
        signature = (slot.shift_template_id, slot.template_code, None)
        idx = column_index.get(signature)
        if idx is None:
            idx = len(columns)
            column_index[signature] = idx
            template = template_by_id.get(slot.shift_template_id) if slot.shift_template_id else None
            columns.append(
                RosterExportColumn(
                    key=f"template-{idx}",
                    title=_template_column_title(template, fallback_slot=slot),
                )
            )
        slot_to_column[slot.id] = idx

    assignments = {assignment.roster_slot_id: assignment for assignment in matrix.assignments}
    members = {member.id: member for member in matrix.team_members}
    entries: dict[tuple[date, int], list[RosterExportCell]] = {}
    for slot in slots_sorted:
        assignment = assignments.get(slot.id)
        if assignment is None:
            continue
        member = members.get(assignment.team_member_id)
        if member is None:
            continue
        col_idx = slot_to_column[slot.id]
        entry_label = _planning_member_label(member)
        entries.setdefault((slot.slot_date, col_idx), []).append(
            RosterExportCell(member_name=entry_label, palette=member_pastel_palette(member.id))
        )

    rows: list[RosterExportRow] = []
    for day in matrix.days:
        row_cells: list[RosterExportCell | None] = []
        for col_idx in range(len(columns)):
            cell_entries = entries.get((day.date, col_idx), [])
            if not cell_entries:
                row_cells.append(None)
                continue
            if len(cell_entries) == 1:
                row_cells.append(cell_entries[0])
                continue
            combined = "\n".join(entry.member_name for entry in cell_entries)
            row_cells.append(RosterExportCell(member_name=combined, palette=None))
        rows.append(RosterExportRow(day=day.date, weekday=_weekday_short(day.weekday), cells=row_cells))

    return RosterExportTable(
        period_label=_period_label(matrix.planning_period.year, matrix.planning_period.month),
        columns=columns,
        rows=rows,
    )


def export_matrix_csv(
    db: Session, planning_period_id: int, *, organization_id: int, shift_group_id: int | None = None
) -> str:
    matrix = get_planning_matrix(db, planning_period_id, organization_id=organization_id, shift_group_id=shift_group_id)
    cells = {(cell.cell_date, cell.team_member_id): cell for cell in matrix.cells}
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["date", *[_planning_member_label(m) for m in matrix.team_members]])
    for day in matrix.days:
        row = [day.date.isoformat()]
        for member in matrix.team_members:
            cell = cells.get((day.date, member.id))
            if cell is None:
                row.append("")
                continue
            value = cell.status
            if cell.comment:
                value = f"{value} - {cell.comment}"
            row.append(value)
        writer.writerow(row)
    return buffer.getvalue()


def export_roster_matrix_csv(
    db: Session, planning_period_id: int, *, organization_id: int, shift_group_id: int | None = None
) -> str:
    matrix = get_roster_matrix(db, planning_period_id, organization_id=organization_id, shift_group_id=shift_group_id)
    slots_by_day: dict[date, list[Any]] = {}
    for slot in matrix.slots:
        slots_by_day.setdefault(slot.slot_date, []).append(slot)
    assignments = {assignment.roster_slot_id: assignment for assignment in matrix.assignments}
    members = {member.id: member for member in matrix.team_members}

    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["date", "slot", "start", "end", "team_member", "template_code", "variant", "category"])
    for day in matrix.days:
        for slot in slots_by_day.get(day.date, []):
            assignment = assignments.get(slot.id) if slot else None
            member = members.get(assignment.team_member_id) if assignment else None
            writer.writerow(
                [
                    day.date.isoformat(),
                    slot.label or "",
                    slot.starts_at.isoformat() if slot.starts_at else "",
                    slot.ends_at.isoformat() if slot.ends_at else "",
                    _planning_member_label(member) if member else "",
                    slot.template_code or "",
                    slot.variant_label or "",
                    slot.category or "",
                ]
            )
    return buffer.getvalue()


def export_roster_matrix_xlsx(
    db: Session, planning_period_id: int, *, organization_id: int, shift_group_id: int | None = None
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX exports") from exc

    table = build_roster_export_table(
        db, planning_period_id, organization_id=organization_id, shift_group_id=shift_group_id
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Roster"

    max_col = max(2, len(table.columns) + 1)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    sheet["A1"] = "Shift Planner"
    sheet["A1"].font = Font(bold=True, size=14, color="16202A")
    sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="3DD6A5")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    sheet["A2"] = f"Roster {table.period_label}"
    sheet["A2"].font = Font(size=11, color="334155")

    header_row = 4
    thin = Side(border_style="thin", color="CBD5E1")
    sheet.cell(row=header_row, column=1, value="Weekday")
    sheet.cell(row=header_row, column=2, value="Date")
    for idx, column in enumerate(table.columns, start=3):
        sheet.cell(row=header_row, column=idx, value=column.title)
    for col_idx in range(1, len(table.columns) + 3):
        header_cell = sheet.cell(row=header_row, column=col_idx)
        header_cell.font = Font(bold=True, color="0F172A")
        header_cell.fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    data_start = header_row + 1
    for row_idx, row in enumerate(table.rows, start=data_start):
        weekday_cell = sheet.cell(row=row_idx, column=1, value=row.weekday)
        weekday_cell.font = Font(bold=True, color="0F172A")
        weekday_cell.fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
        weekday_cell.alignment = Alignment(horizontal="center", vertical="center")
        weekday_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        day_cell = sheet.cell(row=row_idx, column=2, value=row.day.isoformat())
        day_cell.font = Font(bold=True, color="0F172A")
        day_cell.fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
        day_cell.alignment = Alignment(horizontal="center", vertical="center")
        day_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, roster_cell in enumerate(row.cells, start=3):
            cell = sheet.cell(
                row=row_idx,
                column=col_idx,
                value=roster_cell.member_name if roster_cell else "",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if roster_cell:
                cell.fill = PatternFill(fill_type="solid", fgColor=roster_cell.palette.fill_hex[1:])
                cell.font = Font(color=roster_cell.palette.text_hex[1:], bold=True)

    sheet.freeze_panes = "C5"
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 12
    for idx in range(3, len(table.columns) + 3):
        sheet.column_dimensions[sheet.cell(row=header_row, column=idx).column_letter].width = 24

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_roster_matrix_pdf(
    db: Session, planning_period_id: int, *, organization_id: int, shift_group_id: int | None = None
) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise RuntimeError("reportlab is required for PDF exports") from exc

    table = build_roster_export_table_by_template(
        db, planning_period_id, organization_id=organization_id, shift_group_id=shift_group_id
    )
    data: list[list[str]] = [["Wochentag", "Datum", *[column.title for column in table.columns]]]
    for row in table.rows:
        data.append([row.weekday, row.day.strftime("%d.%m.%Y"), *[(cell.member_name if cell else "") for cell in row.cells]])

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        title=f"Roster {table.period_label}",
        leftMargin=24,
        rightMargin=24,
        topMargin=28,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    title = Paragraph(f"<b>Shift Planner</b> - Roster {table.period_label}", styles["Title"])
    subtitle = Paragraph("Final roster matrix export", styles["Normal"])
    story = [title, Spacer(1, 6), subtitle, Spacer(1, 12)]

    table_component = Table(data, repeatRows=1)
    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16202a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 1), (1, -1), colors.HexColor("#f8fafc")),
            ("FONTNAME", (0, 1), (1, -1), "Helvetica-Bold"),
        ]
    )
    for row_idx, row in enumerate(table.rows, start=1):
        for col_idx, cell in enumerate(row.cells, start=2):
            if cell is None or cell.palette is None:
                continue
            style.add("BACKGROUND", (col_idx, row_idx), (col_idx, row_idx), colors.HexColor(cell.palette.fill_hex))
            style.add("TEXTCOLOR", (col_idx, row_idx), (col_idx, row_idx), colors.HexColor(cell.palette.text_hex))
            style.add("FONTNAME", (col_idx, row_idx), (col_idx, row_idx), "Helvetica-Bold")
    table_component.setStyle(style)
    story.append(table_component)
    document.build(story)
    return buffer.getvalue()
